"""xlsx exports (openpyxl) for earnings and life care plan results.

These functions only format the engine's typed results into a workbook. They do
no math. Each returns the path written.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from engine.earnings import EarningsResult
from engine.lcp import LCPResult

_MONEY = "#,##0"
_MONEY2 = "#,##0.00"
_PCT = "0.00%"


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 3


def _sources_sheet(wb, sources: list | None) -> None:
    """Append a 'Sources' sheet with the raw look-up data behind the numbers."""
    if not sources:
        return
    ws = wb.create_sheet("Sources")
    ws.append(["Raw data appendix"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for src in sources:
        ws.append([src.get("title", "")])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        if src.get("citation"):
            ws.append([src["citation"]])
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)
        cols = src.get("columns") or []
        if cols:
            ws.append(list(cols))
            for c in ws[ws.max_row]:
                c.font = Font(bold=True)
        for row in src.get("rows") or []:
            ws.append(list(row))
        ws.append([])
    _autosize(ws)


def earnings_workbook(result: EarningsResult, path: str, *,
                      sources: list | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Earnings"
    headers = ["Year", "Portion", "Gross earnings", "AIF",
               "Adjusted income", "Present value"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in result.rows:
        ws.append([
            r.year, r.portion, r.gross_earnings, r.aif,
            r.adjusted_income, r.present_value,
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = _PCT
        row[2].number_format = _MONEY
        row[3].number_format = _PCT
        row[4].number_format = _MONEY
        row[5].number_format = _MONEY

    ws.append([])
    for label, value in (
        ("Past present value", result.past_present_value),
        ("Future present value", result.future_present_value),
        ("Total present value", result.total_present_value),
    ):
        ws.append([label, None, None, None, None, value])
        ws.cell(row=ws.max_row, column=6).number_format = _MONEY2
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _autosize(ws)
    _sources_sheet(wb, sources)
    wb.save(path)
    return path


def lcp_workbook(result: LCPResult, path: str, *,
                 sources: list | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "LCP items"
    ws.append(["Item", "Category", "Occurrences", "Nominal total",
               "Present value", "Overlaps LHHS"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for it in result.items:
        ws.append([
            it.name, it.category, it.occurrences,
            it.nominal_total, it.present_value,
            "yes" if it.overlaps_household else "",
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = _MONEY
        row[4].number_format = _MONEY
    _autosize(ws)

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Category", "Present value"])
    ws2["A1"].font = ws2["B1"].font = Font(bold=True)
    for cat, pv in result.category_present_value.items():
        ws2.append([cat, pv])
        ws2.cell(row=ws2.max_row, column=2).number_format = _MONEY
    ws2.append([])
    for label, value in (
        ("Lifetime present value", result.lifetime_present_value),
        ("Household-services overlap", result.household_overlap_present_value),
        ("Lifetime excluding overlap", result.lifetime_excluding_overlap()),
    ):
        ws2.append([label, value])
        ws2.cell(row=ws2.max_row, column=2).number_format = _MONEY2
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    _autosize(ws2)

    _sources_sheet(wb, sources)
    wb.save(path)
    return path
